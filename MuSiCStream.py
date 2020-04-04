"""
PyMSATm -- Python Multipurpose Simulation analysis Tools
Package for MuSiC Simulation Software
Copyright (C) 2017  Xumiao Zhou, Li Yang
E-mail: 1026715200@qq.com   liyang@wit.edu.cn
Contact Address :
Li Yang
School of Chemical Engineering and Pharmacy
Wuhan Institute of Technology
No.206, Guanggu 1st road
Donghu New & High Technology Development Zone
Wuhan, Hubei Province, P.R. China
Postcode: 430205
"""

import os
import math
import time
import random
import re
import sys

OpenExcel = False

try:
    from openpyxl import Workbook
    OpenExcel = True
except ImportError:
    Workbook = None
    print("""Warning: please install openpyxl!
This coed will do not generate the Excel file!
It will still generate the txt file!""")


def ReadBasicInfo():
    """Read parameters from GlueParameters file"""

    EquilibriumStep, ProductionStep, HEPCP, HEPCE, Multiple = 10000000, 10000000, 100, 100, 2
    InputPath, OutputPath, AtomParameterPath, TaskSuffix, MaterialInputFormat = '..', '..', '..', '', 'mol'
    GasType, GasAtomTypeNum, GasAtomType, GasPartialPressure, TemperatureList, PressureList,\
        TorqueSetting, MuSiCSetting, Nodes = [], [], [], [], [], [], [], [], ['1:ppn=1']
    CutOff, GridSpacingP, GridSpacingE = 12.8, 2.0, 2.0
    MakeGCMC, UsePmap, UseEmap, UsePost, MakePmap, MakeEmap, MakeTorque, KeyOne, KeyTwo,\
        PDBCharges = False, False, False, False, False, False, False, False, False, False

    with open('GlueParameters', 'r') as File:
        for Line in File.readlines():
            if Line.strip():
                WordList = Line.strip().split()
                if len(WordList) > 1 or KeyOne or KeyTwo:
                    if WordList[0] == '#':
                        continue

                    # Controlled part
                    elif WordList[0] == 'MakeGCMC:' and WordList[1] == 'open':
                        MakeGCMC = True
                    elif WordList[0] == 'UsePmap:' and WordList[1] == 'yes':
                        UsePmap = True
                    elif WordList[0] == 'UseEmap:' and WordList[1] == 'yes':
                        UseEmap = True
                    elif WordList[0] == 'UsePost:' and WordList[1] == 'yes':
                        UsePost = True
                    elif WordList[0] == 'MakePmap:' and WordList[1] == 'open':
                        MakePmap = True
                    elif WordList[0] == 'MakeEmap:' and WordList[1] == 'open':
                        MakeEmap = True
                    elif WordList[0] == 'MakeTorque:' and WordList[1] == 'open':
                        MakeTorque = True
                    elif WordList[0] == 'UseChargesFromPDBFile:' and WordList[1] == 'yes':
                        PDBCharges = True

                    # Basic part
                    elif WordList[0] == 'InputPath:':
                        InputPath = WordList[1]
                    elif WordList[0] == 'MaterialInputFormat:':
                        MaterialInputFormat = WordList[1]
                    elif WordList[0] == 'OutputPath:':
                        OutputPath = WordList[1]
                    elif WordList[0] == 'AtomParameterPath:':
                        AtomParameterPath = WordList[1]
                    elif WordList[0] == 'GasType:':
                        GasType = list(WordList[1:])
                    elif WordList[0] == 'GasAtomTypeNum:':

                        for i in WordList[1:]:
                            GasAtomTypeNum.append(int(i))

                    elif WordList[0] == 'GasAtomType:':
                        GasAtomType = list(WordList[1:])
                    elif WordList[0] == 'Multiple:':
                        Multiple = int(WordList[1])
                    elif WordList[0] == 'CutOff:':
                        CutOff = float(WordList[1])

                    # GCMC part

                    elif WordList[0] == 'GasPartialPressure:':

                        for j in WordList[1:]:
                            GasPartialPressure.append(str(j))

                    elif WordList[0] == 'TemperatureList(K):':

                        for l in WordList[1:]:
                            TemperatureList.append(float(l))

                    elif WordList[0] == 'PressureList(kPa):':

                        for k in WordList[1:]:
                            PressureList.append(float(k))

                    elif WordList[0] == 'EquilibriumStep:':
                        EquilibriumStep = int(WordList[1])
                    elif WordList[0] == 'ProductionStep:':
                        ProductionStep = int(WordList[1])

                    # Pmap part
                    elif WordList[0] == 'GridSpacingP(Ang):':
                        GridSpacingP = float(WordList[1])
                    elif WordList[0] == 'HighEndPotentialCutoffP(kJ/mol):':
                        HEPCP = int(WordList[1])

                    # Emap part
                    elif WordList[0] == 'GridSpacingE(Ang):':
                        GridSpacingE = float(WordList[1])
                    elif WordList[0] == 'HighEndPotentialCutoffE(kJ/mol):':
                        HEPCE = int(WordList[1])

                    # Torque part
                    elif WordList[0] == 'Nodes:':
                        Nodes = WordList[1:]
                    elif WordList[0] == 'TaskSuffix:':
                        TaskSuffix = WordList[1]
                    elif WordList[0] == 'TorqueSetting:':
                        KeyOne = True
                    elif WordList[0] == 'MuSiCSetting:':
                        KeyOne = False
                        KeyTwo = True
                    elif WordList[0] == 'END':
                        KeyTwo = False
                    elif KeyOne:
                        TorqueSetting.append(Line)
                    elif KeyTwo:
                        MuSiCSetting.append(Line)

    return (
        InputPath,
        OutputPath,
        AtomParameterPath,
        MakeTorque,
        GasType,
        GasAtomTypeNum,
        GasAtomType,
        GasPartialPressure,
        TemperatureList,
        PressureList,
        CutOff,
        MakeGCMC,
        UsePmap,
        UseEmap,
        UsePost,
        MakePmap,
        MakeEmap,
        EquilibriumStep,
        ProductionStep,
        GridSpacingP,
        HEPCP,
        GridSpacingE,
        HEPCE,
        Multiple,
        TorqueSetting,
        MuSiCSetting,
        Nodes,
        TaskSuffix,
        PDBCharges,
        MaterialInputFormat)


def ReadAtomParameter(AtomParameterPath):
    """read the file of atom parameter and return the parameter dictionary"""

    AtomParameter = os.path.join(AtomParameterPath, 'AtomParameter')

    Key1, Key2, Key3 = False, False, False
    MaterialAtomDictionary, GasAtomDictionary, MassDictionary = {}, {}, {}
    SpecialPair, SpecialPairList = [], []

    with open(AtomParameter, 'r') as File:
        for Line in File.readlines():
            if Line.strip():
                WordList = Line.strip().split()
                if WordList[0] == '#':
                    continue
                elif WordList[0] == 'MaterialAtom:':
                    Key1 = True
                elif WordList[0] == 'GasAtom:':
                    Key1 = False
                    Key2 = True
                elif WordList[0] == 'SpecialPair:':
                    Key2 = False
                    Key3 = True

                # MaterialAtom
                elif Key1 and WordList[0] != 'Number':
                    MaterialAtomDictionary[WordList[1]] = WordList[2:4]
                    MassDictionary[WordList[1]] = WordList[5]
                elif Key2 and WordList[0] != 'Number':
                    GasAtomDictionary[WordList[1]] = WordList[2:4]
                    MassDictionary[WordList[1]] = WordList[4]
                elif Key3 and WordList[0] != 'Number':
                    SpecialPair.append(WordList[1:3])
                    SpecialPair.append(WordList[3:5])

            SpecialPairList.append(SpecialPair)

    return MaterialAtomDictionary, GasAtomDictionary, SpecialPairList, MassDictionary


def ReadMaterialNameList(InputPath, MaterialInputFormat):
    """find mol file and return a available name list in input path"""

    FullNameList = os.listdir(InputPath)
    MaterialPathList = []

    for FullName in FullNameList:
        Name, NameType = os.path.splitext(FullName)
        if NameType == '.' + MaterialInputFormat:
            MaterialPath = os.path.join(InputPath, FullName)
            MaterialPathList.append(MaterialPath)

    MaterialPathList.sort(key=str.lower)

    if len(MaterialPathList) < 1:
        print(
            'There are not have ' +
            MaterialInputFormat +
            ' file in the specified folder')
        exit()

    return MaterialPathList


def ReadMaterialInfoAndMakeMaterialsMolFiles(
        OutputPath,
        MaterialPathList,
        CutOff,
        Multiple,
        PDBCharges,
        MaterialInputFormat):
    """read every material and return a list which have every material with full information"""

    MaterialInfoList = []
    ID = 0
    Time = time.strftime('%Y-%m-%d', time.localtime(time.time()))

    if os.path.exists(os.path.join(OutputPath, 'Mols')):
        pass
    else:
        os.mkdir(os.path.join(OutputPath, 'Mols'))

    for Range, MaterialPath in enumerate(MaterialPathList):

        AtomNum = 0
        MaterialsCharge = 0.0
        MaterialInfo, ElementList, CellLength, CellAngle, AtomInfo, AtomInfoList, CellSize = [
        ], [], [], [], [], [], []

        MaterialName = os.path.splitext(MaterialPath.split('/')[-1])[0]
        MaterialInfo.append(MaterialName)
        MaterialNewName = re.sub(r'\W', '', MaterialName)

        if len(MaterialNewName) > 20:
            ID += 1
            MaterialNewName = MaterialNewName[:20]
            MaterialNewName = MaterialNewName + '_ID_' + str(ID)

        if MaterialInputFormat == 'pdb':
            with open(MaterialPath, 'r') as File:
                with open(os.path.join(OutputPath, 'Mols', MaterialNewName) + '.mol', 'w') as Mol:

                    for Line in File.readlines():
                        if Line[0:6] == 'CRYST1':
                            CellLength.append(
                                float('%.3f' % (float(Line[6:15].strip()))))
                            CellLength.append(
                                float('%.3f' % (float(Line[15:24].strip()))))
                            CellLength.append(
                                float('%.3f' % (float(Line[24:33].strip()))))
                            CellAngle.append(
                                float('%.2f' % (float(Line[33:40].strip()))))
                            CellAngle.append(
                                float('%.2f' % (float(Line[40:47].strip()))))
                            CellAngle.append(
                                float('%.2f' % (float(Line[47:54].strip()))))
                        elif Line[0:6] == 'ATOM  ':
                            AtomNum += 1
                            AtomInfo = []
                            AtomInfo.append(int(Line[6:11].strip()))  # serial
                            AtomInfo.append(
                                str(Line[12:16].strip()))  # atom name
                            AtomInfo.append(float(Line[30:38].strip()))  # x
                            AtomInfo.append(float(Line[38:46].strip()))  # y
                            AtomInfo.append(float(Line[46:54].strip()))  # z

                            if PDBCharges:
                                AtomInfo.append(
                                    float(Line[54:60].strip()))  # charge
                                MaterialsCharge += float(Line[54:60].strip())
                            else:
                                AtomInfo.append(float('0.0'))  # charge

                            # atom name list
                            ElementList.append(str(Line[12:16].strip()))
                            AtomInfoList.append(AtomInfo)

                    if abs(MaterialsCharge) < 0.001:
                        Mol.write(
                            '# Basic Molecule Information\n# Created by PyMSATm at %s\n'
                            'Molecule_name: %s  \n\nCoord_Info: Listed Cartesian None\n' %
                            (Time, MaterialNewName))
                    else:
                        Mol.write(
                            '# Basic Molecule Information\n# Created by PyMSATm at %s\n'
                            'Molecule_name: %s  CHARGED\n\nCoord_Info: Listed Cartesian None\n' %
                            (Time, MaterialNewName))

                    Mol.write('   %d\n' % (AtomNum))

                    for AtomInfo2 in AtomInfoList:
                        Mol.write(
                            '{:<6d} {:>8} {:>8} {:>8} {:>6} {:>6}   0   0\n'.format(
                                AtomInfo2[0],
                                AtomInfo2[2],
                                AtomInfo2[3],
                                AtomInfo2[4],
                                AtomInfo2[1],
                                AtomInfo2[5]))

                    Mol.write(
                        '\nFundcell_Info:  Listed\n%.5f  %.5f  %.5f\n%.5f  %.5f  %.5f\n0.00000  0.00000  0.00000\n%.5f  %.5f  %.5f' %
                        (CellLength[0],
                         CellLength[1],
                            CellLength[2],
                            CellAngle[0],
                            CellAngle[1],
                            CellAngle[2],
                            CellLength[0],
                            CellLength[1],
                            CellLength[2]))

        if MaterialInputFormat == 'mol':

            KeyOne, KeyTwo, KeyThree = False, False, False

            with open(MaterialPath, 'r') as File:
                with open(os.path.join(OutputPath, 'Mols', MaterialNewName) + '.mol', 'w') as Mol:

                    for Line in File.readlines():
                        if Line.strip():  # skip blank
                            WordList = Line.strip().split()
                            if WordList[0] == 'Coord_Info:':
                                KeyOne = True
                            elif WordList[0] == 'Fundcell_Info:':
                                KeyOne = False
                                KeyTwo = True
                            elif len(WordList) > 1 and KeyOne:
                                AtomNum += 1
                                AtomInfo = []
                                AtomInfo.append(
                                    int(WordList[0].strip()))  # serial
                                AtomInfo.append(
                                    str(WordList[4].strip()))  # atom name
                                AtomInfo.append(
                                    float(WordList[1].strip()))  # x
                                AtomInfo.append(
                                    float(WordList[2].strip()))  # y
                                AtomInfo.append(
                                    float(WordList[3].strip()))  # z
                                AtomInfo.append(
                                    float(WordList[5].strip()))  # charge
                                # atom name list
                                ElementList.append(str(WordList[4].strip()))
                                AtomInfoList.append(AtomInfo)
                                MaterialsCharge += float(WordList[5].strip())
                            elif len(WordList) > 1 and KeyTwo:
                                KeyTwo = False
                                KeyThree = True
                                CellLength.append(
                                    float('%.3f' % (float(WordList[0].strip()))))
                                CellLength.append(
                                    float('%.3f' % (float(WordList[1].strip()))))
                                CellLength.append(
                                    float('%.3f' % (float(WordList[2].strip()))))
                            elif len(WordList) > 1 and KeyThree:
                                KeyThree = False
                                CellAngle.append(
                                    float('%.2f' % (float(WordList[0].strip()))))
                                CellAngle.append(
                                    float('%.2f' % (float(WordList[1].strip()))))
                                CellAngle.append(
                                    float('%.2f' % (float(WordList[2].strip()))))

                    if abs(MaterialsCharge) < 0.001:
                        Mol.write(
                            '# Basic Molecule Information\n# Created by PyMSATm at %s\n'
                            'Molecule_name: %s  \n\nCoord_Info: Listed Cartesian None\n' %
                            (Time, MaterialNewName))
                    else:
                        Mol.write(
                            '# Basic Molecule Information\n# Created by PyMSATm at %s\n'
                            'Molecule_name: %s  CHARGED\n\nCoord_Info: Listed Cartesian None\n' %
                            (Time, MaterialNewName))

                    Mol.write('   %d\n' % (AtomNum))

                    for AtomInfo2 in AtomInfoList:
                        Mol.write(
                            '{:<6d} {:>8.4f} {:>8.4f} {:>8.4f} {:>8s} {:>10.6f}   0   0\n'.format(
                                AtomInfo2[0],
                                AtomInfo2[2],
                                AtomInfo2[3],
                                AtomInfo2[4],
                                AtomInfo2[1],
                                AtomInfo2[5]))

                    Mol.write(
                        '\nFundcell_Info:  Listed\n%.4f  %.4f  %.4f\n%.4f  %.4f  %.4f\n0.00000  0.00000  0.00000\n%.4f  %.4f  %.4f' %
                        (CellLength[0],
                         CellLength[1],
                            CellLength[2],
                            CellAngle[0],
                            CellAngle[1],
                            CellAngle[2],
                            CellLength[0],
                            CellLength[1],
                            CellLength[2]))

        if CellAngle[0] == 90.0 and CellAngle[1] == 90.0 and CellAngle[2] == 90.0:
            Orthogonality = True
        else:
            Orthogonality = False

        if Orthogonality:
            for Length in CellLength:
                Count = 1
                while Length * Count < CutOff * Multiple + 1:
                    Count += 1
                CellSize.append(str(Count))
        else:    # definite the lengths of x, y, z in orthogonal directions

            SinA, SinB, SinC = math.sin(
                math.radians(
                    CellAngle[0])), math.sin(
                math.radians(
                    CellAngle[1])), math.sin(
                math.radians(
                    CellAngle[2]))
            CosA, CosB, CosC = math.cos(
                math.radians(
                    CellAngle[0])), math.cos(
                math.radians(
                    CellAngle[1])), math.cos(
                math.radians(
                    CellAngle[2]))
            CosParameterA = (CosC * CosA - CosB) / (SinC * SinA)
            SinParameterA = math.sqrt(1 - CosParameterA ** 2)
            CosParameterB = (CosA * CosB - CosC) / (SinA * SinB)
            SinParameterB = math.sqrt(1 - CosParameterB ** 2)
            CosParameterC = (CosB * CosC - CosA) / (SinB * SinC)
            SinParameterC = math.sqrt(1 - CosParameterC ** 2)
            x = CellLength[0] * SinC * SinParameterA
            y = CellLength[1] * SinA * SinParameterB
            z = CellLength[2] * SinB * SinParameterC
            TransLength = [x, y, z]

            for Length in TransLength:
                Count = 1
                while float(Length) * Count < CutOff * Multiple + 1:
                    Count += 1
                CellSize.append(str(Count))

        MaterialInfo.append(AtomNum)
        MaterialInfo.append(CellLength)
        MaterialInfo.append(CellAngle)
        MaterialInfo.append(CellSize)
        MaterialInfo.append(list(set(ElementList)))
        MaterialInfo.append(Orthogonality)
        MaterialInfo.append(MaterialNewName)
        MaterialInfoList.append(MaterialInfo)

    return MaterialInfoList


def MakeStatisticsResults(MaterialInfoList, OutputPath, OpenExcel):
    """count materials information and output information"""

    TotElementList, TotElementNameList = [], []
    Statistics = os.path.join(OutputPath, "Statistics")

    if os.path.exists(Statistics):
        pass
    else:
        os.makedirs(Statistics)

    with open(os.path.join(Statistics, "MaterialsInfo.txt"), 'w') as Information:
        Information.write(
            'Data: %s\nMaterials                                           Mol_ID                                              AtomsNumber       CellAngle                   CellLength                   SuperCell     Elements\n' %
            (time.strftime(
                '%Y-%m-%d',
                time.localtime(
                    time.time()))))

        for Range, MaterialInfo in enumerate(MaterialInfoList):

            AngleList, LengthList, ElementList, ElementNameList = [], [], [], []

            for Length in MaterialInfo[2]:
                LengthList.append('%s' % ('%.4f' % (Length)))

            for Angle in MaterialInfo[3]:
                AngleList.append('%s' % ('%.4f' % (Angle)))

            for Element in MaterialInfo[5]:
                ElementList.append(Element.split('_')[0])
                ElementNameList.append(Element)

            Information.write(
                '%-50s  %-50s  %-16s  %-28s%-29s%-14s%-20s\n' %
                (MaterialInfo[0],
                 MaterialInfo[7],
                 MaterialInfo[1],
                 ','.join(AngleList),
                    ','.join(LengthList),
                    ','.join(
                    MaterialInfo[4]),
                    ','.join(ElementList)))

            TotElementList.extend(ElementList)
            TotElementNameList.extend(ElementNameList)

        TotElementList = sorted(set(TotElementList))
        TotElementNameList = sorted(set(TotElementNameList))
        Information.write('\nTotElements: %s' % (','.join(TotElementList)))

    if OpenExcel:

        MaterialsInfo = Workbook()
        worksheet = MaterialsInfo.create_sheet('Information', 0)
        worksheet.cell(row=1, column=1, value='Data:%s' %
                       (time.strftime('%Y-%m-%d', time.localtime(time.time()))))
        worksheet.cell(row=2, column=1, value='Materials')
        worksheet.cell(row=2, column=2, value='Mol_ID')
        worksheet.cell(row=2, column=3, value='AtomsNumber')
        worksheet.cell(row=2, column=4, value='CellAngle')
        worksheet.cell(row=2, column=5, value='CellLength')
        worksheet.cell(row=2, column=6, value='SuperCell')
        worksheet.cell(row=2, column=7, value='Elements')

        for Range, MaterialInfo in enumerate(MaterialInfoList):

            AngleList, LengthList, ElementList = [], [], []

            for Length in MaterialInfo[2]:
                LengthList.append('%s' % ('%.4f' % (Length)))

            for Angle in MaterialInfo[3]:
                AngleList.append('%s' % ('%.4f' % (Angle)))

            for Element in MaterialInfo[5]:
                ElementList.append(Element.split('_')[0])

            worksheet.cell(row=3 + Range, column=1, value=MaterialInfo[0])
            worksheet.cell(row=3 + Range, column=2, value=MaterialInfo[7])
            worksheet.cell(row=3 + Range, column=3, value=int(MaterialInfo[1]))
            worksheet.cell(row=3 + Range, column=4, value=' '.join(LengthList))
            worksheet.cell(row=3 + Range, column=5, value=' '.join(AngleList))
            worksheet.cell(
                row=3 + Range,
                column=6,
                value=' '.join(
                    MaterialInfo[4]))
            worksheet.cell(
                row=3 + Range,
                column=7,
                value=' '.join(ElementList))

        worksheet.cell(
            row=3 +
            len(MaterialInfoList),
            column=1,
            value='TotElements: %s' %
            (' '.join(TotElementList)))
        MaterialsInfo.save(os.path.join(Statistics, 'MaterialsInfo.xlsx'))

    return TotElementNameList


def MakeAtomFileAndMap(
        OutputPath,
        TotElementNameList,
        GasAtomType,
        MassDictionary):
    """make xxx.atm file"""

    AtomOutputPath = '%s/Atoms' % (OutputPath)
    MapOutputPath = '%s/Maps' % (OutputPath)

    if os.path.exists(AtomOutputPath):
        pass
    else:
        os.makedirs(AtomOutputPath)

    if os.path.exists(MapOutputPath):
        pass
    else:
        os.makedirs(MapOutputPath)

    for i in TotElementNameList:
        with open('%s/%s.atm' % (AtomOutputPath, i), 'w') as AtomFile:
            AtomFile.write('''# Basic Atom Information
Atom_Name:           %s
Atom_Symbol:         %s
Atom_SS_Charge:      0
Atom_SZ_Charge:      0
Atom_Mass:           %s
Atom_Valency:        0''' % (i, i.split('_')[0], MassDictionary.get(i)))

    for j in GasAtomType:
        with open('%s/%s.atm' % (AtomOutputPath, j), 'w') as AtomFile:
            AtomFile.write('''# Basic Atom Information
Atom_Name:           %s
Atom_Symbol:         %s
Atom_SS_Charge:      0
Atom_SZ_Charge:      0
Atom_Mass:           %s
Atom_Valency:        0''' % (j, j.split('_')[0], MassDictionary.get(j)))

    with open('%s/Probe.atm' % (AtomOutputPath), 'w') as AtomFile:
        AtomFile.write('''# Basic Atom Information
Atom_Name:           Probe
Atom_Symbol:         P
Atom_SS_Charge:      0
Atom_SZ_Charge:      0
Atom_Mass:           0
Atom_Valency:        0''')


def MakeGasMolFile(OutputPath, GasAtomType, GasAtomDictionary, MakeEmap):
    """make xxx.mol file"""
    MolOutputPath = '%s/Mols' % (OutputPath)
    for i in GasAtomType:
        if float(GasAtomDictionary.get(i)[0]) != 0.0:
            with open('%s/%s.mol' % (MolOutputPath, i), 'w') as MolFile:
                MolFile.write('''# Basic Molecule Information
Molecule_Name: %s

Coord_Info: Listed Cartesian Rigid
    1     # number of atoms in molecule
 1  0.000  0.000  0.0000  %s   0    0   0 # x,y,z,name,charge,set,type

Molecule_DOF: 3''' % (i, i))

    if MakeEmap:
        with open(os.path.join(MolOutputPath, "Probe.mol"), 'w') as MolFile:
            MolFile.write(
                '''# Basic Molecule Information\nMolecule_Name: Probe CHARGED\n
Coord_Info: Listed Cartesian Rigid
1     # number of atoms in molecule
1  0.000  0.000  0.0000  Probe   1    0   0 # x,y,z,name,charge,set,type

Molecule_DOF: 3''')


def MakeGCMCProgram(
        MaterialInfoList,
        TemperatureList,
        PressureList,
        OutputPath,
        MakeTorque,
        GasType,
        GasAtomTypeNum,
        GasAtomType,
        GasPartialPressure,
        CutOff,
        MaterialAtomDictionary,
        GasAtomDictionary,
        SpecialPairList,
        UseEmap,
        UsePmap,
        UsePost,
        EquilibriumStep,
        ProductionStep,
        TorqueSetting,
        MuSiCSetting,
        Nodes,
        TaskSuffix):
    """make mainly file of GCMC Program"""

    def MakeAtomAtomFile(
            GCMCOutputPath,
            MaterialInfo,
            GasAtomType,
            SpecialPairList,
            GasAtomDictionary,
            MaterialAtomDictionary):

        with open('%s/atom_atom_file' % (GCMCOutputPath), 'w') as AtomAtomFile:

            AtomAtomFile.write('-'.center(80, '-'))
            AtomAtomFile.write('\n')

            for i in range(len(MaterialInfo[5])):
                for j in range(len(MaterialInfo[5])):
                    if i <= j:
                        AtomAtomFile.write(
                            '%-10s%-10sOFF\n' %
                            (MaterialInfo[5][i], MaterialInfo[5][j]))

            for k in range(len(GasAtomType)):
                for l in range(len(GasAtomType)):
                    if k <= l:
                        Key = False
                        for SpecialPair in SpecialPairList:
                            if GasAtomType[k] in SpecialPair[0] and GasAtomType[l] in SpecialPair[0] and GasAtomType[k] != GasAtomType[l]:
                                Key = True
                        if not Key:
                            num1 = GasAtomDictionary.get(GasAtomType[k])
                            num2 = GasAtomDictionary.get(GasAtomType[l])
                            sig1 = str('%.3f' %
                                       ((float(num1[0]) + float(num2[0])) / 2))
                            eps1 = str('%.3f' %
                                       ((float(num1[1]) * float(num2[1])) ** 0.5))
                            AtomAtomFile.write(
                                '%-10s%-10s%-10sSIG@%-20sEPS@%-20sHICUT@%-10sLOCUT@0.1000\n%-10s%-10s%-10sHICUT@%-10sALPHA@0.10\n' %
                                (GasAtomType[k],
                                 GasAtomType[l],
                                    'LJ',
                                    sig1,
                                    eps1,
                                    CutOff,
                                    GasAtomType[k],
                                    GasAtomType[l],
                                    'WFCOUL',
                                    CutOff))

            for h in range(len(GasAtomType)):
                for g in range(len(MaterialInfo[5])):
                    Key = False
                    for SpecialPair in SpecialPairList:
                        if GasAtomType[h] in SpecialPair[0] and MaterialInfo[5][g] in SpecialPair[0]:
                            Key = True
                    if not Key:
                        num3 = GasAtomDictionary.get(GasAtomType[h])
                        num4 = MaterialAtomDictionary.get(MaterialInfo[5][g])
                        sig2 = str('%.3f' %
                                   ((float(num3[0]) + float(num4[0])) / 2))
                        eps2 = str('%.3f' %
                                   ((float(num3[1]) * float(num4[1])) ** 0.5))
                        AtomAtomFile.write(
                            '%-10s%-10s%-10sSIG@%-20sEPS@%-20sHICUT@%-10sLOCUT@0.1000\n%-10s%-10s%-10sHICUT@%-10sALPHA@0.10\n' %
                            (GasAtomType[h],
                             MaterialInfo[5][g],
                                'LJ',
                                sig2,
                                eps2,
                                CutOff,
                                GasAtomType[h],
                                MaterialInfo[5][g],
                                'WFCOUL',
                                CutOff))

            for m in SpecialPairList:
                AtomAtomFile.write(
                    '%-10s%-10s%-10sSIG@%-20sEPS@%-20sHICUT@%-10sLOCUT@0.1000\n%-10s%-10s%-10sHICUT@%-10sALPHA@0.10\n' %
                    (m[0][0], m[0][1], 'LJ', m[1][0], m[1][1], CutOff, m[0][0], m[0][1], 'WFCOUL', CutOff))

            AtomAtomFile.write('-'.center(80, '-'))

    def MakeIntramolecularFile(GCMCOutputPath, MaterialInfo, GasType):

        with open('%s/intramolecular_file' % (GCMCOutputPath), 'w') as IntraFile:
            IntraFile.write('Intra:  %s' % (MaterialInfo[7]))
            for i in GasType:
                IntraFile.write('\nIntra:  %s' % (i))

    def MakeMoleMoleFile(
            GCMCOutputPath,
            MaterialInfo,
            GasType,
            UsePmap,
            UseEmap,
            GasAtomTypeNum,
            GasAtomType,
            GasAtomDictionary):

        with open('%s/mole_mole_file' % (GCMCOutputPath), 'w') as MoleMole:
            MoleMole.write(
                '%s  %s  NCOUL  OFF\n%s  %s  COUL  OFF' %
                (MaterialInfo[7], MaterialInfo[7], MaterialInfo[7], MaterialInfo[7]))

            for i in range(len(GasType)):
                for j in range(len(GasType)):
                    if i <= j:
                        MoleMole.write(
                            '\n\n%s  %s  NCOUL  BASIC   LJ  FAST\n%s  %s  COUL  BASIC  FAST  WFCOUL' %
                            (GasType[i], GasType[j], GasType[i], GasType[j]))

            MoleMole.write('\n')

            if UsePmap and MaterialInfo[6]:
                Key = 0
                for k in range(len(GasType)):
                    MoleMole.write(
                        '\n%s  %s  NCOUL  MAP@%s FAST ' %
                        (GasType[k], MaterialInfo[7], MaterialInfo[7]))
                    for l in GasAtomType[Key:Key + GasAtomTypeNum[i]]:
                        pseudo = l.split('_')
                        if pseudo[0] != 'M' and GasAtomDictionary.get(l)[
                                0] != '0':
                            MoleMole.write(
                                '%s@PMAP@%s_in_%s.pmap  ' %
                                (l, l, MaterialInfo[7]))
                    Key = Key + GasAtomTypeNum[i]
            else:
                for k in GasType:
                    MoleMole.write(
                        '\n%s  %s  NCOUL  BASIC   LJ  FAST' %
                        (k, MaterialInfo[7]))

            MoleMole.write('\n')

            if UseEmap and MaterialInfo[6]:
                for m in GasType:
                    MoleMole.write(
                        '\n%s  %s  COUL   MAP@%s FAST all@EMAP@%s_all.emap' %
                        (m, MaterialInfo[7], MaterialInfo[7], MaterialInfo[7]))
            else:
                for m in GasType:
                    MoleMole.write(
                        '\n%s  %s  COUL   BASIC  FAST  WFCOUL' %
                        (m, MaterialInfo[7]))

    def MakeEquilibriumGCMC(
            GCMCOutputPath,
            GasType,
            MaterialInfo,
            EquilibriumStep,
            GasAtomType,
            Temperature,
            Pressure,
            PartialList):

        with open('%s/equilibrium_gcmc.ctr' % (GCMCOutputPath), 'w') as EquilibriumGCMC:
            EquilibriumGCMC.write('''------ General Information ------
%s molecule in %s
%s               # No. of iterations
10000                  # No. of steps between writes to output/log file
10000                  # No. of steps between writes to crash file
100000                 # No. of steps between writes to config. file
1                     # Start numbering simulations from .
30728                 # Iseed
3                     # specifies contents of config file,  (3) only nmoles, nrg ,pair nrg and coords written
%s_in_%s_res          # Restart File to write to
%s_in_%s_con          # Configuration File

------ Atomic Types ------
%s                    # number of atomic types         ''' % (' '.join(GasType), MaterialInfo[7], EquilibriumStep,
                                                              '_'.join(GasType), MaterialInfo[7], '_'.join(GasType),
                                                              MaterialInfo[7], len(MaterialInfo[5]) + len(GasAtomType)))

            for i in GasAtomType:
                EquilibriumGCMC.write('\n\n%s\n%s.atm' % (i, i))

            for j in MaterialInfo[5]:
                EquilibriumGCMC.write('\n\n%s\n%s.atm' % (j, j))

            EquilibriumGCMC.write('''\n------ Molecule Types ------
%s                    # number of molecular types''' % (len(GasType) + 1))

            for k in GasType:
                EquilibriumGCMC.write('\n\n%s\n%s.mol' % (k, k))

            EquilibriumGCMC.write('''\n\n%s               # sorbate
%s.mol                # sorbate coordiCates file
------ Simulation Cell Information ------
%s                    # Fundamental cell file
%s                    # No. of unit cells in x, y, z direction
1, 1, 1               # (1 = Periodic) in x, y, z
------ Forcefield Information ------
BASIC
MOL
atom_atom_file       # atom-atom interaction file
mole_mole_file       # sorbate-sorbate interaction file
intramolecular_file  # intramolecular interaction file/specification
------ Ideal Parameters ------
Ideal                # Equation of State
%s                   # no. of sorbates''' % (MaterialInfo[7], MaterialInfo[7], MaterialInfo[7], ', '.join(MaterialInfo[4]), len(GasType)))

            for l in GasType:
                EquilibriumGCMC.write('\n%s' % (l))

            EquilibriumGCMC.write('''\n------ GCMC Information -------
1                    # No. of iterations
%s                   # temperature
Ideal Parameters     # Tag for the equation of state (NULL = Ideal Gas)
1                    # No. of simulation points
5000                 # Block size for statistics
%s                   # no. of sorbates''' % (Temperature, len(GasType)))

            for m in range(len(GasType)):
                EquilibriumGCMC.write('''\n-------
%s            # Sorbate Came
%s          #  pressure(kpa)
Null               # sitemap fileCame (Null = no sitemap)
4                  # no of gcmc movetypes
2.0, 2.0, 1.0, 1.0   # move type weights
RINSERT              # type of move
RDELETE             # type of move
RTRANSLATE          # type of move
0.2, 1              # Delta Translate, adjust delta option (0=NO, 1=YES)
RROTATE
0.2, 1              # Delta Translate, adjust delta option (0=NO, 1=YES)''' % (GasType[m], PartialList[m] * Pressure))

            EquilibriumGCMC.write(
                '\n------ Configuration Initialization ------')

            for n in GasType:
                EquilibriumGCMC.write(
                    '\n%s             # Sorbate_Type\nGCMC NULL' %
                    (n))

            EquilibriumGCMC.write('''\n%s              # Sorbent_Type
FIXED NULL
------  Main Datafile Information ------
Energy, position, pair_energy  # contents of datafile''' % (MaterialInfo[7]))

    def MakeProductionGCMC(
            GCMCOutputPath,
            GasType,
            MaterialInfo,
            ProductionStep,
            GasAtomType,
            Temperature,
            Pressure,
            PartialList):

        with open('%s/production_gcmc.ctr' % (GCMCOutputPath), 'w') as ProductionGCMC:
            ProductionGCMC.write('''------ General Information ------
%s molecule in %s
%s               # No. of iterations
5000                  # No. of steps between writes to output/log file
5000                  # No. of steps between writes to crash file
20000                 # No. of steps between writes to config. file
2                     # Start numbering simulations from .
30728                 # Iseed
3                     # specifies contents of config file,  (3) only nmoles, nrg ,pair nrg and coords written
%s_in_%s_res          # Restart File to write to
%s_in_%s_con          # Configuration File

------ Atomic Types ------
%s                    # number of atomic types         ''' % (' '.join(GasType), MaterialInfo[7], ProductionStep,
                                                              '_'.join(GasType), MaterialInfo[7], '_'.join(GasType),
                                                              MaterialInfo[7], len(MaterialInfo[5]) + len(GasAtomType)))

            for i in GasAtomType:
                ProductionGCMC.write('\n\n%s\n%s.atm' % (i, i))

            for j in MaterialInfo[5]:
                ProductionGCMC.write('\n\n%s\n%s.atm' % (j, j))

            ProductionGCMC.write('''\n------ Molecule Types ------
%s                    # number of molecular types''' % (len(GasType) + 1))

            for k in GasType:
                ProductionGCMC.write('\n\n%s\n%s.mol' % (k, k))

            ProductionGCMC.write('''\n\n%s               # sorbate
%s.mol                # sorbate coordiCates file
------ Simulation Cell Information ------
%s                    # Fundamental cell file
%s                    # No. of unit cells in x, y, z direction
1, 1, 1               # (1 = Periodic) in x, y, z
------ Forcefield Information ------
BASIC
MOL
atom_atom_file       # atom-atom interaction file
mole_mole_file       # sorbate-sorbate interaction file
intramolecular_file  # intramolecular interaction file/specification
------ Ideal Parameters ------
Ideal                # Equation of State
%s                   # no. of sorbates''' % (MaterialInfo[7], MaterialInfo[7], MaterialInfo[7], ', '.join(MaterialInfo[4]), len(GasType)))

            for l in GasType:
                ProductionGCMC.write('\n%s' % (l))

            ProductionGCMC.write('''\n------ GCMC Information -------
1                    # No. of iterations
%s                   # temperature
Ideal Parameters     # Tag for the equation of state (NULL = Ideal Gas)
1                    # No. of simulation points
1000                 # Block size for statistics
%s                   # no. of sorbates''' % (Temperature, len(GasType)))

            for m in range(len(GasType)):
                ProductionGCMC.write('''\n-------
%s            # Sorbate Came
%s          #  pressure(kpa)
Null               # sitemap fileCame (Null = no sitemap)
4                  # no of gcmc movetypes
2.0, 2.0, 1.0, 1.0   # move type weights
RINSERT              # type of move
RDELETE             # type of move
RTRANSLATE          # type of move
0.2, 1              # Delta Translate, adjust delta option (0=NO, 1=YES)
RROTATE
0.2, 1              # Delta Translate, adjust delta option (0=NO, 1=YES)''' % (GasType[m], PartialList[m] * Pressure))

            ProductionGCMC.write(
                '\n------ Configuration Initialization ------')

            for n in GasType:
                ProductionGCMC.write(
                    '\n%s             # Sorbate_Type\nRESTARTFILE  %s_in_%s_res.1' %
                    (n, '_'.join(GasType), MaterialInfo[7]))

            ProductionGCMC.write('''\n%s              # Sorbent_Type
RESTARTFILE  %s_in_%s_res.1
------  Main Datafile Information ------
Energy, position, pair_energy  # contents of datafile''' % (MaterialInfo[7], '_'.join(GasType), MaterialInfo[7]))

    def MakePostFile(GCMCOutputPath, GasType, MaterialInfo):

        with open('%s/post.ctr' % (GCMCOutputPath), 'w') as Post:
            Post.write('''------ Post Processor Information ------
GCMC                      # Type of simlation GCMC, MD
%s_in_%s_con              # basename for config files what your .con files are called in your gcmc folder
2                         # The first and last numbers of the .con files that were created in the gcmc folder
regenerated_post          # name for new ctrlfile that will be regenerated
result_post               # Base name for output files
20, 0                      # Percentages of data to skipped at start and end

------ Post : Energy Average Info -------
20
------ Post : Loading Average Info --------
20
''' % ('_'.join(GasType), MaterialInfo[7]))

    def MakeTorqueFile(
            GCMCOutputPath,
            Nodes,
            TaskSuffix,
            TorqueSetting,
            MuSiCSetting,
            UsePost,
            OutputPath):

        Node = random.choice(Nodes)

        with open('%s/run_gcmc.pbs' % (GCMCOutputPath), 'w') as Torque:
            Torque.write('''#!/bin/bash
#PBS -l nodes=%s
#PBS -N MuSiC_gcmc.%s
#PBS -o music_gcmc_jobs.out
#PBS -j oe

#
# The number of processors you desire is indicated by replacing
# <nproc> above.
#

#
# GROMACS path and arguments to mdrun :
#
cd $PBS_O_WORKDIR

# =============== Environment Setting ============================ #\n''' % (Node, TaskSuffix))

            for i in TorqueSetting:
                Torque.write('%s' % (i))

            Torque.write(
                '''# =============== Don't Change Above Setting ===================== #

echo "============The computed nodes============"
cp    -f  $PBS_NODEFILE  NODE.txt
echo  "User:               " $USER
cat   $PBS_NODEFILE
echo  "Job ID:             " $PBS_JOBID
echo  "Job Cookie:         " $PBS_JOBCOOKIE
echo  "Using executable:   " `which mpirun`
echo  `date`
echo "============Finished setting=============="

# =========== Setting Jobs ============================ #\n''')

            for j in MuSiCSetting:
                Torque.write('%s' % (j))

            Torque.write(
                '''export ATOMSDIR=%s
export MOLSDIR=%s
export PMAPDIR=%s
export EMAPDIR=%s
export SMAPDIR=%s''' %
                (os.path.join(
                    OutputPath, 'Atoms'), os.path.join(
                    OutputPath, 'Mols'), os.path.join(
                    OutputPath, 'Maps'), os.path.join(
                    OutputPath, 'Maps'), os.path.join(
                        OutputPath, 'Maps')))

            Torque.write('''
# =========== Setting Jobs ============================ #

# +++++++++++++++ Start Computing +++++++++++++++++++++ #

TIME_DIR=$(date '+%Y-%m-%d_%H-%M-%S')
TIME_DIR="${USER}_jobs_${TIME_DIR}_${PBS_JOBID}"
if [ -d /utmp ]; then
    TEMP_DIR=/utmp/${USER}/${TIME_DIR}
else
    TEMP_DIR=/temp/${USER}/${TIME_DIR}
fi
mkdir -p ${TEMP_DIR}
cp -rf * ${TEMP_DIR}
cd ${TEMP_DIR}
rm -f music_gcmc_jobs.out
echo "The temp direcotry: " ${TEMP_DIR}
echo "============Finished setting=============="

echo "+++++++++++++ Run MuSic ++++++++++++++++++++++++++++"
music_gcmc  equilibrium_gcmc.ctr >  equilibrium_gcmc.txt
echo `date`
music_gcmc  production_gcmc.ctr >  production_gcmc.txt
echo `date`''')

            if UsePost:
                Torque.write('\nmusic_post  post.ctr >  post.txt\necho `date`')

            Torque.write(
                '''\necho "+++++++++++++ Finish MuSic +++++++++++++++++++++++++"

cd $PBS_O_WORKDIR
cp -rf ${TEMP_DIR}/* .
rm -rf ${TEMP_DIR}


echo "All files were copied back!"
echo "The work direcotry: " $PBS_O_WORKDIR
echo `date`
echo "============Finished Job =============="''')

    def main():

        for MaterialInfo in MaterialInfoList:
            for Partial in GasPartialPressure:
                for Temperature in TemperatureList:
                    for Pressure in PressureList:
                        GCMCOutputPath = '%s/%s/%s/%s/%s/%sK/%skPa' % (OutputPath, 'GCMC', '_'.join(
                            GasType), MaterialInfo[7], Partial, Temperature, Pressure)

                        if os.path.exists(GCMCOutputPath):
                            pass
                        else:
                            os.makedirs(GCMCOutputPath)
                        PartialList = [float(x) for x in Partial.split('_')]
                        MakeAtomAtomFile(
                            GCMCOutputPath,
                            MaterialInfo,
                            GasAtomType,
                            SpecialPairList,
                            GasAtomDictionary,
                            MaterialAtomDictionary)
                        MakeIntramolecularFile(
                            GCMCOutputPath, MaterialInfo, GasType)
                        MakeMoleMoleFile(
                            GCMCOutputPath,
                            MaterialInfo,
                            GasType,
                            UsePmap,
                            UseEmap,
                            GasAtomTypeNum,
                            GasAtomType,
                            GasAtomDictionary)
                        MakeEquilibriumGCMC(
                            GCMCOutputPath,
                            GasType,
                            MaterialInfo,
                            EquilibriumStep,
                            GasAtomType,
                            Temperature,
                            Pressure,
                            PartialList)
                        MakeProductionGCMC(
                            GCMCOutputPath,
                            GasType,
                            MaterialInfo,
                            ProductionStep,
                            GasAtomType,
                            Temperature,
                            Pressure,
                            PartialList)

                        if MakeTorque:
                            MakeTorqueFile(
                                GCMCOutputPath,
                                Nodes,
                                TaskSuffix,
                                TorqueSetting,
                                MuSiCSetting,
                                UsePost,
                                OutputPath)
                        if UsePost:
                            MakePostFile(GCMCOutputPath, GasType, MaterialInfo)

    if __name__ == '__main__':
        main()


def MakePmapProgram(
        MaterialInfoList,
        OutputPath,
        GasType,
        GasAtomType,
        SpecialPairList,
        GasAtomDictionary,
        MaterialAtomDictionary,
        GridSpacingP,
        HEPCP,
        CutOff,
        Nodes,
        TaskSuffix,
        TorqueSetting,
        MuSiCSetting):
    """make mainly file of Pmap Prograam"""

    def MakeAtomAtomFile(
            PmapOutputPath,
            MaterialInfo,
            GasAtomType,
            SpecialPairList,
            GasAtomDictionary,
            MaterialAtomDictionary,
            CutOff):

        with open('%s/atom_atom_file' % (PmapOutputPath), 'w') as AtomAtomFile:

            AtomAtomFile.write('-'.center(80, '-'))
            AtomAtomFile.write('\n')

            for i in range(len(MaterialInfo[5])):
                for j in range(len(MaterialInfo[5])):
                    if i <= j:
                        AtomAtomFile.write(
                            '%-10s%-10sOFF\n' %
                            (MaterialInfo[5][i], MaterialInfo[5][j]))

            for k in range(len(GasAtomType)):
                for l in range(len(GasAtomType)):
                    if k <= l:
                        Key = False
                        for SpecialPair in SpecialPairList:
                            if GasAtomType[k] in SpecialPair[0] and GasAtomType[l] in SpecialPair[0] and GasAtomType[k] != GasAtomType[l]:
                                Key = True
                        if not Key:
                            num1 = GasAtomDictionary.get(GasAtomType[k])
                            num2 = GasAtomDictionary.get(GasAtomType[l])
                            sig1 = str('%.3f' %
                                       ((float(num1[0]) + float(num2[0])) / 2))
                            eps1 = str('%.3f' %
                                       ((float(num1[1]) * float(num2[1])) ** 0.5))
                            AtomAtomFile.write(
                                '%-10s%-10s%-10sSIG@%-20sEPS@%-20sHICUT@%-10sLOCUT@0.1000\n%-10s%-10s%-10sHICUT@%-10sALPHA@0.10\n' %
                                (GasAtomType[k],
                                 GasAtomType[l],
                                    'LJ',
                                    sig1,
                                    eps1,
                                    CutOff,
                                    GasAtomType[k],
                                    GasAtomType[l],
                                    'WFCOUL',
                                    CutOff))

            for h in range(len(GasAtomType)):
                for g in range(len(MaterialInfo[5])):
                    Key = False
                    for SpecialPair in SpecialPairList:
                        if GasAtomType[h] in SpecialPair[0] and MaterialInfo[5][g] in SpecialPair[0]:
                            Key = True
                    if not Key:
                        num3 = GasAtomDictionary.get(GasAtomType[h])
                        num4 = MaterialAtomDictionary.get(MaterialInfo[5][g])
                        sig2 = str('%.3f' %
                                   ((float(num3[0]) + float(num4[0])) / 2))
                        eps2 = str('%.3f' %
                                   ((float(num3[1]) * float(num4[1])) ** 0.5))
                        AtomAtomFile.write(
                            '%-10s%-10s%-10sSIG@%-20sEPS@%-20sHICUT@%-10sLOCUT@0.1000\n%-10s%-10s%-10sHICUT@%-10sALPHA@0.10\n' %
                            (GasAtomType[h],
                             MaterialInfo[5][g],
                                'LJ',
                                sig2,
                                eps2,
                                CutOff,
                                GasAtomType[h],
                                MaterialInfo[5][g],
                                'WFCOUL',
                                CutOff))

            for m in SpecialPairList:
                AtomAtomFile.write(
                    '%-10s%-10s%-10sSIG@%-20sEPS@%-20sHICUT@%-10sLOCUT@0.1000\n%-10s%-10s%-10sHICUT@%-10sALPHA@0.10\n' %
                    (m[0][0], m[0][1], 'LJ', m[1][0], m[1][1], CutOff, m[0][0], m[0][1], 'WFCOUL', CutOff))

            AtomAtomFile.write('-'.center(80, '-'))

    def MakeIntramolecularFile(
            PmapOutputPath,
            MaterialInfo,
            GasAtomType,
            GasAtomDictionary):

        with open('%s/intramolecular_file' % (PmapOutputPath), 'w') as IntraFile:
            IntraFile.write('Intra:  %s' % (MaterialInfo[7]))
            for i in GasAtomType:
                pseudo = i.split('_')
                if pseudo[0] != 'M' and GasAtomDictionary.get(i)[0] != '0':
                    IntraFile.write('\nIntra:  %s' % (i))

    def MakeMoleMolePmapFile(
            PmapOutputPath,
            MaterialInfo,
            GasAtomType,
            GasAtomDictionary):

        with open('%s/mole_mole_pmap_file' % (PmapOutputPath), 'w') as MoleMolePmap:
            MoleMolePmap.write('''%s %s NCOUL OFF
%s %s COUL OFF\n\n''' % (MaterialInfo[7], MaterialInfo[7], MaterialInfo[7], MaterialInfo[7]))

            for i in GasAtomType:
                pseudo = i.split('_')
                if pseudo[0] != 'M' and GasAtomDictionary.get(i)[0] != '0':
                    MoleMolePmap.write('''%s %s NCOUL OFF
%s %s COUL OFF

%s %s NCOUL BASIC LJ FAST
%s %s COUL OFF\n\n''' % (i, i, i, i, i, MaterialInfo[7], i, MaterialInfo[7]))

    def MakePmapMaker(
            PmapOutputPath,
            MaterialInfo,
            GasAtomType,
            GridSpacingP,
            HEPCP,
            GasAtomDictionary):

        for i in GasAtomType:
            pseudo = i.split('_')
            if pseudo[0] != 'M' and GasAtomDictionary.get(i)[0] != '0':
                with open('%s/pmap_maker_%s_in_%s.ctr' % (PmapOutputPath, i, MaterialInfo[7]), 'w') as PmapMaker:
                    PmapMaker.write(
                        '''------ General Information ------------------------------------------
%s molecule in %s
1                                        # No. of iterations
1                                        # No. of steps between writes to output/log file
2                                        # No. of steps between writes to crash file
2                                        # No. of steps between writes to config. file
1                                        # Start numbering simulations from .
30728                                    # Iseed
1                                        # specifies contents of config file
%s_in_%s.res                             # Restart File to write to
%s_in_%s.con                             # Configuration File

------ Atomic Types --------------------------------------------------
%s                                       # number of atomic types

%s
%s.atm''' %
                        (i, MaterialInfo[7], i, MaterialInfo[7], i, MaterialInfo[7], len(
                            MaterialInfo[5]) + 1, i, i))

                    for j in MaterialInfo[5]:
                        PmapMaker.write('\n\n%s\n%s.atm' % (j, j))

                    PmapMaker.write(
                        '''\n------ Molecule Types -------------------------------------------------
2

%s
%s.mol

%s
%s.mol
------ Simulation Cell Information ------------------------------------
%s                                   # Fundamental cell file
%s                                   # No. of unit cells in x, y, z direction
1, 1, 1                              # (1 = Periodic) in x, y, z
------ Forcefield Information -------------------------------------------
BASIC
MOL
atom_atom_file                       # atom-atom interaction file
mole_mole_pmap_file                  # sorbate-sorbate interaction file
intramolecular_file                  # intramolecular interaction file/specification
------ Mapmaker Information -----------------------------------------------
1                                    # Number of maps to make

%s                           # Sorbent to map
%s                                 # Sorbate to probe map with
NCOUL LJ                             # Interaction type to map
%s                                  # Approxiamte grid spacing (Ang)
%s                                # High end potential cutoff (kJ/mol)
%s_in_%s.pmap              # Map filename or AUTO
------ Configuration Initialization -------------------------------------
%s                                 # Sorbate_Type
MOLECULE NULL
%s                           # Sorbate_Type
FIXED NULL''' %
                        (i,
                         i,
                         MaterialInfo[7],
                            MaterialInfo[7],
                            MaterialInfo[7],
                            ', '.join(
                             MaterialInfo[4]),
                            MaterialInfo[7],
                            i,
                            GridSpacingP,
                            HEPCP,
                            i,
                            MaterialInfo[7],
                            i,
                            MaterialInfo[7]))

    def MakeTorqueFile(
            PmapOutputPath,
            Nodes,
            TaskSuffix,
            TorqueSetting,
            MuSiCSetting,
            GasAtomType,
            GasAtomDictionary,
            MaterialInfo,
            OutputPath):

        Node = random.choice(Nodes)

        with open('%s/run_pmapmaker.pbs' % (PmapOutputPath), 'w') as Torque:
            Torque.write('''#!/bin/bash
#PBS -l nodes=%s
#PBS -N MuSiC_pmap.%s
#PBS -o music_pmap_jobs.out
#PBS -j oe

#
# The number of processors you desire is indicated by replacing
# <nproc> above.
#

#
# GROMACS path and arguments to mdrun :
#
cd $PBS_O_WORKDIR

# =============== Environment Setting ============================ #\n''' % (Node, TaskSuffix))

            for i in TorqueSetting:
                Torque.write('%s' % (i))

            Torque.write(
                '''# =============== Don't Change Above Setting ===================== #

echo "============The computed nodes============"
cp    -f  $PBS_NODEFILE  NODE.txt
echo  "User:               " $USER
cat   $PBS_NODEFILE
echo  "Job ID:             " $PBS_JOBID
echo  "Job Cookie:         " $PBS_JOBCOOKIE
echo  "Using executable:   " `which mpirun`
echo  `date`
echo "============Finished setting=============="

# =========== Setting Jobs ============================ #\n''')

            for j in MuSiCSetting:
                Torque.write('%s' % (j))

            Torque.write(
                '''export ATOMSDIR=%s
            export MOLSDIR=%s
            export PMAPDIR=%s
            export EMAPDIR=%s
            export SMAPDIR=%s''' %
                (os.path.join(
                    OutputPath, 'Atoms'), os.path.join(
                    OutputPath, 'Mols'), os.path.join(
                    OutputPath, 'Maps'), os.path.join(
                    OutputPath, 'Maps'), os.path.join(
                        OutputPath, 'Maps')))

            Torque.write(
                '''# =========== Setting Jobs ============================ #

# +++++++++++++++ Start Computing +++++++++++++++++++++ #

TIME_DIR=$(date '+%Y-%m-%d_%H-%M-%S')
TIME_DIR="${USER}_jobs_${TIME_DIR}_${PBS_JOBID}"
if [ -d /utmp ]; then
    TEMP_DIR=/utmp/${USER}/${TIME_DIR}
else
    TEMP_DIR=/temp/${USER}/${TIME_DIR}
fi
mkdir -p ${TEMP_DIR}
cp -rf * ${TEMP_DIR}
cd ${TEMP_DIR}
rm -f music_pmap_jobs.out
echo "The temp direcotry: " ${TEMP_DIR}
echo "============Finished setting=============="

echo "+++++++++++++ Run MuSic ++++++++++++++++++++++++++++"\n''')

            for i in GasAtomType:
                pseudo = i.split('_')
                if pseudo[0] != 'M' and GasAtomDictionary.get(i)[0] != '0':
                    Torque.write(
                        'music_mapmaker  pmap_maker_%s_in_%s.ctr >  pmap_maker_%s_in_%s.txt\necho `date`\n' %
                        (i, MaterialInfo[7], i, MaterialInfo[7]))

            Torque.write(
                '''echo "+++++++++++++ Finish MuSic +++++++++++++++++++++++++"

cd $PBS_O_WORKDIR
cp -rf ${TEMP_DIR}/* .
rm -rf ${TEMP_DIR}


echo "All files were copied back!"
echo "The work direcotry: " $PBS_O_WORKDIR
echo `date`
echo "============Finished Job =============="''')

    def main():

        for MaterialInfo in MaterialInfoList:
            if MaterialInfo[6]:
                PmapOutputPath = '%s/%s/%s/%s' % (OutputPath,
                                                  'MakePmap',
                                                  '_'.join(GasType),
                                                  MaterialInfo[7])
                if os.path.exists(PmapOutputPath):
                    pass
                else:
                    os.makedirs(PmapOutputPath)

                MakeAtomAtomFile(
                    PmapOutputPath,
                    MaterialInfo,
                    GasAtomType,
                    SpecialPairList,
                    GasAtomDictionary,
                    MaterialAtomDictionary,
                    CutOff)
                MakeMoleMolePmapFile(
                    PmapOutputPath,
                    MaterialInfo,
                    GasAtomType,
                    GasAtomDictionary)
                MakePmapMaker(
                    PmapOutputPath,
                    MaterialInfo,
                    GasAtomType,
                    GridSpacingP,
                    HEPCP,
                    GasAtomDictionary)
                MakeIntramolecularFile(
                    PmapOutputPath,
                    MaterialInfo,
                    GasAtomType,
                    GasAtomDictionary)
                MakeTorqueFile(
                    PmapOutputPath,
                    Nodes,
                    TaskSuffix,
                    TorqueSetting,
                    MuSiCSetting,
                    GasAtomType,
                    GasAtomDictionary,
                    MaterialInfo,
                    OutputPath)

    if __name__ == '__main__':
        main()


def MakeEmapProgram(
        MaterialInfoList,
        OutputPath,
        GridSpacingE,
        HEPCE,
        Nodes,
        TaskSuffix,
        TorqueSetting,
        MuSiCSetting):
    """make mainly file of Emap Prograam"""

    def MakeIntramolecularFile(EmapOutputPath, MaterialInfo):

        with open('%s/intramolecular_file' % (EmapOutputPath), 'w') as IntraFile:
            IntraFile.write('Intra:  %s\nIntra: Probe' % (MaterialInfo[7]))

    def MakeSpcSpcFileEmap(EmapOutputPath, MaterialInfo):

        with open('%s/spc_spc_file_emap' % (EmapOutputPath), 'w') as SpcFile:
            SpcFile.write('''%s    %s NCOUL OFF
%s    %s  COUL   OFF
Probe    %s  NCOUL OFF
Probe    %s  COUL  SUM FAST FIXED EWALD SFACTOR KMAX@15 KAPPA@6.7 LOCUT@1e-10
Probe    Probe NCOUL OFF
Probe    Probe  COUL   OFF''' % (MaterialInfo[7], MaterialInfo[7], MaterialInfo[7], MaterialInfo[7], MaterialInfo[7], MaterialInfo[7]))

    def MakeEmapMaker(EmapOutputPath, MaterialInfo, GridSpacingE, HEPCE):

        with open('%s/emap_maker.ctr' % (EmapOutputPath), 'w') as EmapMaker:

            EmapMaker.write(
                '''------ General Information ------------------------------------------
Probe in %s
1                           # No. of iterations
1                           # No. of steps between writes to output/log file
1                           # No. of steps between writes to crash file
1                           # No. of steps between writes to config. file
2                           # Start numbering simulations from .
030728                      # Iseeed
4                           # specifies contents of config file,
%s_all.res          # Restart File to write to
%s_all.con          # Configuration File
------ Atomic Types --------------------------------------------------
%s                           # number of atomic types

Probe
Probe.atm''' %
                (MaterialInfo[7], MaterialInfo[7], MaterialInfo[7], len(
                    MaterialInfo[5]) + 1))

            for i in MaterialInfo[5]:
                EmapMaker.write('\n\n%s\n%s.atm' % (i, i))

            EmapMaker.write(
                '''\n------ Molecule Types -------------------------------------------------
2                             # number of sorbate types

Probe                         # sorbate
Probe.mol                     # sorbate coordinates file

%s                    # sorbate
%s.mol                # sorbate coordinates file
------ Simulation Cell Information --------------------------------------
%s                    # Fundamental cell type
%s                       # No. of unit cells in x, y, z direction
1, 1, 1                       # (1 = Periodic) in x, y, z
------ Forcefield Information -------------------------------------------
BASIC
MOL
atom_atom_file                # atom-atom interaction file
spc_spc_file_emap             # sorbate-sorbate interaction file
intramolecular_file           # intramolecular interaction file/specification
------ Mapmaker Information --------------------------------------------
1                             # Number of maps to make

%s                    # Sorbate to map
Probe                         # Sorbate to probe map with
COUL EWALD                    # Interaction type to map
%s                           # Approximate grid spacing (Ang)
%s                         # High end potential cutoff (kJ/mol)
%s_all.emap           # Map filename or AUTO
------ Configuration Initialization -------------------------------------
Probe                         # Sorbate_Type
Molecule NULL                 # Source Filename
%s                    # Sorbate_Type
Fixed NULL                    # Source Filename''' %
                (MaterialInfo[7],
                 MaterialInfo[7],
                 MaterialInfo[7],
                 ', '.join(
                    MaterialInfo[4]),
                    MaterialInfo[7],
                    GridSpacingE,
                    HEPCE,
                    MaterialInfo[7],
                    MaterialInfo[7]))

    def MakeTorqueFile(
            EmapOutputPath,
            Nodes,
            TaskSuffix,
            TorqueSetting,
            MuSiCSetting,
            OutputPath):

        Node = random.choice(Nodes)

        with open('%s/run_emapmaker.pbs' % (EmapOutputPath), 'w') as Torque:
            Torque.write('''#!/bin/bash
#PBS -l nodes=%s
#PBS -N MuSiC_emap.%s
#PBS -o music_emap_jobs.out
#PBS -j oe

#
# The number of processors you desire is indicated by replacing
# <nproc> above.
#

#
# GROMACS path and arguments to mdrun :
#
cd $PBS_O_WORKDIR

# =============== Environment Setting ============================ #\n''' % (Node, TaskSuffix))

            for i in TorqueSetting:
                Torque.write('%s' % (i))

            Torque.write(
                '''# =============== Don't Change Above Setting ===================== #

echo "============The computed nodes============"
cp    -f  $PBS_NODEFILE  NODE.txt
echo  "User:               " $USER
cat   $PBS_NODEFILE
echo  "Job ID:             " $PBS_JOBID
echo  "Job Cookie:         " $PBS_JOBCOOKIE
echo  "Using executable:   " `which mpirun`
echo  `date`
echo "============Finished setting=============="

# =========== Setting Jobs ============================ #\n''')

            for j in MuSiCSetting:
                Torque.write('%s' % (j))

            Torque.write(
                '''export ATOMSDIR=%s
export MOLSDIR=%s
export PMAPDIR=%s
export EMAPDIR=%s
export SMAPDIR=%s\n''' %
                (os.path.join(
                    OutputPath, 'Atoms'), os.path.join(
                    OutputPath, 'Mols'), os.path.join(
                    OutputPath, 'Maps'), os.path.join(
                    OutputPath, 'Maps'), os.path.join(
                        OutputPath, 'Maps')))

            Torque.write(
                '''# =========== Setting Jobs ============================ #

# +++++++++++++++ Start Computing +++++++++++++++++++++ #

TIME_DIR=$(date '+%Y-%m-%d_%H-%M-%S')
TIME_DIR="${USER}_jobs_${TIME_DIR}_${PBS_JOBID}"
if [ -d /utmp ]; then
    TEMP_DIR=/utmp/${USER}/${TIME_DIR}
else
    TEMP_DIR=/temp/${USER}/${TIME_DIR}
fi
mkdir -p ${TEMP_DIR}
cp -rf * ${TEMP_DIR}
cd ${TEMP_DIR}
rm -f music_emap_jobs.out
echo "The temp direcotry: " ${TEMP_DIR}
echo "============Finished setting=============="

echo "+++++++++++++ Run MuSic ++++++++++++++++++++++++++++"
music_mapmaker  emap_maker.ctr >  emap_maker.txt
echo `date`
echo "+++++++++++++ Finish MuSic +++++++++++++++++++++++++"

cd $PBS_O_WORKDIR
cp -rf ${TEMP_DIR}/* .
rm -rf ${TEMP_DIR}


echo "All files were copied back!"
echo "The work direcotry: " $PBS_O_WORKDIR
echo `date`
echo "============Finished Job =============="''')

    def main():

        for MaterialInfo in MaterialInfoList:
            if MaterialInfo[6]:
                EmapOutputPath = '%s/%s/%s' % (OutputPath,
                                               'MakeEmap', MaterialInfo[7])
                if os.path.exists(EmapOutputPath):
                    pass
                else:
                    os.makedirs(EmapOutputPath)

                MakeSpcSpcFileEmap(EmapOutputPath, MaterialInfo)
                MakeEmapMaker(
                    EmapOutputPath,
                    MaterialInfo,
                    GridSpacingE,
                    HEPCE)
                MakeIntramolecularFile(EmapOutputPath, MaterialInfo)
                MakeTorqueFile(
                    EmapOutputPath,
                    Nodes,
                    TaskSuffix,
                    TorqueSetting,
                    MuSiCSetting,
                    OutputPath)

    if __name__ == '__main__':
        main()


def main():
    """main function"""

    Ready = input("Are you ready to run programme? (y/n):")
    if Ready == "y" or Ready == "Y" or Ready == "yes" or Ready == "Yes" or Ready == "YES":
        pass
    else:
        sys.exit()
    print('Code is executing ......')

    InputPath, OutputPath, AtomParameterPath, MakeTorque, GasType, GasAtomTypeNum, GasAtomType, GasPartialPressure, \
        TemperatureList, PressureList, CutOff, MakeGCMC, UsePmap, UseEmap, UsePost, MakePmap, MakeEmap, EquilibriumStep, \
        ProductionStep, GridSpacingP, HEPCP, GridSpacingE, HEPCE, Multiple, TorqueSetting, MuSiCSetting, Nodes, \
        TaskSuffix, PDBCharges, MaterialInputFormat = ReadBasicInfo()

    # MaterialAtomDictionary: 'Zr_m': ['2.783', '34.751'] SpecialPairList:
    # [['C_co2', 'O_co2'], ['3.25', '27']]
    MaterialAtomDictionary, GasAtomDictionary, SpecialPairList, MassDictionary = ReadAtomParameter(
        AtomParameterPath)

    MaterialPathList = ReadMaterialNameList(InputPath, MaterialInputFormat)

    MaterialInfoList = ReadMaterialInfoAndMakeMaterialsMolFiles(
        OutputPath, MaterialPathList, CutOff, Multiple, PDBCharges, MaterialInputFormat)
    # MaterialInfoList: ['cd3btb2', '99', [10.4595, 14.3920, 14.7634], [69.6080, 82.4760, 88.0430], ['4', '3', '3'], ['H_m', 'Cd_m', 'O_m', 'C_m'], False, 'cd3btb2']

    TotElementNameList = MakeStatisticsResults(
        MaterialInfoList, OutputPath, OpenExcel)

    MakeAtomFileAndMap(
        OutputPath,
        TotElementNameList,
        GasAtomType,
        MassDictionary)

    if MakeGCMC:
        MakeGCMCProgram(
            MaterialInfoList,
            TemperatureList,
            PressureList,
            OutputPath,
            MakeTorque,
            GasType,
            GasAtomTypeNum,
            GasAtomType,
            GasPartialPressure,
            CutOff,
            MaterialAtomDictionary,
            GasAtomDictionary,
            SpecialPairList,
            UseEmap,
            UsePmap,
            UsePost,
            EquilibriumStep,
            ProductionStep,
            TorqueSetting,
            MuSiCSetting,
            Nodes,
            TaskSuffix)

    if MakePmap:
        MakePmapProgram(
            MaterialInfoList,
            OutputPath,
            GasType,
            GasAtomType,
            SpecialPairList,
            GasAtomDictionary,
            MaterialAtomDictionary,
            GridSpacingP,
            HEPCP,
            CutOff,
            Nodes,
            TaskSuffix,
            TorqueSetting,
            MuSiCSetting)
        MakeGasMolFile(OutputPath, GasAtomType, GasAtomDictionary, MakeEmap)

    if MakeEmap:
        MakeEmapProgram(
            MaterialInfoList,
            OutputPath,
            GridSpacingE,
            HEPCE,
            Nodes,
            TaskSuffix,
            TorqueSetting,
            MuSiCSetting)

    print("Completed!")


if __name__ == '__main__':
    main()
