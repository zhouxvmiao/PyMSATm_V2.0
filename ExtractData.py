##########################################################################
# PyMSATm -- Python Multipurpose Simulation analysis Tools Package for MuSiC    #
# Simulation Software                                                           #
# Copyright (C) 2017  Xumiao Zhou, Li Yang                                      #
#                                                                               #
# 1026715200@qq.com                                                             #
# liyang@wit.edu.cn                                                             #
#                                                                               #
# Contact Address :                                                             #
# Li Yang                                                                       #
# School of Chemical Engineering and Pharmacy                                   #
# Wuhan Institute of Technology                                                 #
# No.206, Guanggu 1st road                                                      #
# Donghu New & High Technology Development Zone                                 #
# Wuhan, Hubei Province, P.R. China                                             #
# Postcode: 430205                                                              #
##########################################################################

import os

OpenExcel = False

try:
    from openpyxl import Workbook
    OpenExcel = True
except BaseException:
    print("""Warning: please install openpyxl!
This coed will do not generate the Excel file!
It will still generate the txt file!""")

def ReadBasicInfo():
    """Read parameters from GlueParameters file"""

    ReadingPath = '..'
    GasType, TemperatureList, PressureList, GasPartialPressure = [], [], [], []
    ExtractEnergyData = False

    with open('GlueParameters', 'r') as File:
        for Line in File.readlines():
            if Line.strip():
                WordList = Line.strip().split()
                if len(WordList) > 1:
                    if WordList[0] == '#':
                        continue
                    elif WordList[0] == 'ExtractEnergyData:' and WordList[1] == 'yes':
                        ExtractEnergyData = True
                    elif WordList[0] == 'OutputPath:':
                        ReadingPath = WordList[1]
                    elif WordList[0] == 'GasType:':
                        GasType = list(WordList[1:])

                    elif WordList[0] == 'GasPartialPressure:':

                        for j in WordList[1:]:
                            GasPartialPressure.append(str(j))

                    elif WordList[0] == 'TemperatureList(K):':

                        for l in WordList[1:]:
                            TemperatureList.append(float(l))

                    elif WordList[0] == 'PressureList(kPa):':

                        for k in WordList[1:]:
                            PressureList.append(float(k))

    return (
        ReadingPath,
        GasType,
        TemperatureList,
        PressureList,
        GasPartialPressure,
        ExtractEnergyData)


def ExtractData(
        ReadingPath,
        GasType,
        TemperatureList,
        PressureList,
        ExtractEnergyData,
        GasPartialPressure):
    """return the extracting data as a multiple dictionaries"""

    HalfReadingPath = os.path.join(
        ReadingPath, 'GCMC', '%s' %
        ('_'.join(GasType)))
    MaterialsNameList = os.listdir(HalfReadingPath)
    MaterialsListLoadingDictionary = {}
    MaterialsListEnergyDictionary = {}

    for MaterialsName in MaterialsNameList:
        PartialListLoadingDictionary = {}
        PartialListEnergyDictionary = {}
        for Partial in GasPartialPressure:
            TemperatureListLoadingDictionary = {}
            TemperatureListEnergyDictionary = {}
            for Temperature in TemperatureList:
                PressureListLoadingDictionary = {}
                PressureListEnergyDictionary = {}
                for Pressure in PressureList:
                    GasListLoadingDictionary = {}
                    GasListEnergyDictionary = {}
                    FullReadingPath = os.path.join(
                        HalfReadingPath,
                        str(MaterialsName),
                        Partial,
                        str(Temperature) + 'K',
                        str(Pressure) + 'kPa')

                    with open(os.path.join(FullReadingPath, 'production_gcmc.txt'), 'rb') as CollectionFile:
                        CollectionFile.seek(-10000, 2)
                        for Line in CollectionFile.readlines():
                            if Line.strip():
                                WordList = Line.strip().split()
                                if len(WordList) > 1:
                                    if WordList[1] == b'P' and WordList[4] == b'loading':
                                        GasName = str(
                                            WordList[0].decode().strip(':'))
                                        GasListLoadingDictionary[GasName] = str(
                                            WordList[6].decode())

                    if ExtractEnergyData:
                        for Gas in GasType:
                            GasListEnergyDictionary[Gas] = 0.0
                        with open(os.path.join(FullReadingPath, 'result_post'), 'r') as PostFile:
                            for Line in PostFile.readlines():
                                if Line.strip():
                                    WordList = Line.strip().split()
                                    if WordList[1] == 'Coulombic' or WordList[1] == 'NonCoulom':
                                        if '--' in WordList[0]:
                                            InterName = WordList[0].split('--')
                                            if InterName[0] in GasType:
                                                GasListEnergyDictionary[InterName[0]] += float(
                                                    WordList[5])
                                            if InterName[1] in GasType:
                                                GasListEnergyDictionary[InterName[0]] += float(
                                                    WordList[5])

                    PressureListLoadingDictionary[Pressure] = GasListLoadingDictionary
                    PressureListEnergyDictionary[Pressure] = GasListEnergyDictionary
                TemperatureListLoadingDictionary[Temperature] = PressureListLoadingDictionary
                TemperatureListEnergyDictionary[Temperature] = PressureListEnergyDictionary
            PartialListLoadingDictionary[Partial] = TemperatureListLoadingDictionary
            PartialListEnergyDictionary[Partial] = TemperatureListEnergyDictionary
        MaterialsListLoadingDictionary[MaterialsName] = PartialListLoadingDictionary
        MaterialsListEnergyDictionary[MaterialsName] = PartialListEnergyDictionary

    return (
        MaterialsListLoadingDictionary,
        MaterialsListEnergyDictionary,
        MaterialsNameList)


def OutputTxt(
        ReadingPath,
        ExtractEnergyData,
        TemperatureList,
        PressureList,
        MaterialsNameList,
        GasType,
        MaterialsListLoadingDictionary,
        MaterialsListEnergyDictionary,
        GasPartialPressure):

    ExtractDataOutputPath = os.path.join(ReadingPath, 'Results')
    LoadingExtractDataOutputPath = os.path.join(
        ExtractDataOutputPath, 'Loading')
    EnergyExtractDataOutputPath = os.path.join(ExtractDataOutputPath, 'Energy')
    MaterialsNameList.sort()
    for Partial in GasPartialPressure:
        for Gas in GasType:
            for Temperature in TemperatureList:

                if os.path.exists(
                    os.path.join(
                        LoadingExtractDataOutputPath,
                        Partial,
                        Gas)):
                    pass
                else:
                    os.makedirs(
                        os.path.join(
                            LoadingExtractDataOutputPath,
                            Partial,
                            Gas))

                with open(os.path.join(LoadingExtractDataOutputPath, Partial, Gas, str(Temperature) + 'K' + '.txt'), 'w') as Txt:
                    Txt.write('MaterialsName                ')
                    for Pressure in PressureList:
                        Txt.write('%-15s' % (str(Pressure) + 'kPa'))
                    Txt.write('\n\n')
                    for Material in MaterialsNameList:
                        Txt.write('%-30s' % (Material))
                        for Pressure in PressureList:
                            Txt.write(
                                '%-15.4f' %
                                (float(
                                    MaterialsListLoadingDictionary[Material][Partial][Temperature][Pressure][Gas])))
                        Txt.write('\n')

    if ExtractEnergyData:
        for Partial in GasPartialPressure:
            for Gas in GasType:
                for Temperature in TemperatureList:

                    if os.path.exists(
                        os.path.join(
                            EnergyExtractDataOutputPath,
                            Partial,
                            Gas)):
                        pass
                    else:
                        os.makedirs(
                            os.path.join(
                                EnergyExtractDataOutputPath,
                                Partial,
                                Gas))

                    with open(os.path.join(EnergyExtractDataOutputPath, Partial, Gas, str(Temperature) + 'K' + '.txt'), 'w') as Txt:
                        Txt.write('MaterialsName                 ')
                        for Pressure in PressureList:
                            Txt.write('%-15s' % (str(Pressure) + 'kPa'))
                        Txt.write('\n\n')
                        for Material in MaterialsNameList:
                            Txt.write('%-30s' % (Material))
                            for Pressure in PressureList:
                                Txt.write(
                                    '%-15.4f' %
                                    (float(
                                        MaterialsListEnergyDictionary[Material][Partial][Temperature][Pressure][Gas])))
                            Txt.write('\n')


def OutputExcel(
        ReadingPath,
        ExtractEnergyData,
        TemperatureList,
        PressureList,
        MaterialsNameList,
        GasType,
        MaterialsListLoadingDictionary,
        MaterialsListEnergyDictionary,
        GasPartialPressure):

    ExtractDataOutputPath = os.path.join(ReadingPath, 'Results')
    LoadingExtractDataOutputPath = os.path.join(
        ExtractDataOutputPath, 'Loading')
    EnergyExtractDataOutputPath = os.path.join(ExtractDataOutputPath, 'Energy')
    GasNum = len(GasType)
    Loading = Workbook()
    MaterialsNameList.sort()
    for Partial in GasPartialPressure:
        if os.path.exists(os.path.join(LoadingExtractDataOutputPath, Partial)):
            pass
        else:
            os.makedirs(os.path.join(LoadingExtractDataOutputPath, Partial))
        for Temperature in TemperatureList:
            worksheet = Loading.create_sheet(str(Temperature), 0)

            for PressureSequence, Pressure in enumerate(PressureList):
                worksheet.cell(
                    row=1,
                    column=2 +
                    PressureSequence *
                    GasNum,
                    value=Pressure)
                for GasSequence, Gas in enumerate(GasType):
                    worksheet.cell(
                        row=2,
                        column=2 +
                        PressureSequence *
                        GasNum +
                        GasSequence,
                        value=Gas)

            for MaterialsSequence, Material in enumerate(MaterialsNameList):
                worksheet.cell(
                    row=3 + MaterialsSequence,
                    column=1,
                    value=Material)
                for PressureSequence, Pressure in enumerate(PressureList):
                    for GasSequence, Gas in enumerate(GasType):
                        worksheet.cell(
                            row=3 + MaterialsSequence,
                            column=2 + PressureSequence * GasNum + GasSequence,
                            value=float(
                                MaterialsListLoadingDictionary[Material][Partial][Temperature][Pressure][Gas]))
        Loading.save(
            os.path.join(
                ExtractDataOutputPath,
                '%s.xlsx' %
                (Partial)))

    if ExtractEnergyData:
        Energy = Workbook()

        for Partial in GasPartialPressure:
            if os.path.exists(
                os.path.join(
                    EnergyExtractDataOutputPath,
                    Partial)):
                pass
            else:
                os.makedirs(os.path.join(EnergyExtractDataOutputPath, Partial))
            for Temperature in TemperatureList:
                worksheet = Energy.create_sheet(str(Temperature), 0)

                for PressureSequence, Pressure in enumerate(PressureList):
                    worksheet.cell(
                        row=1,
                        column=2 +
                        PressureSequence *
                        GasNum,
                        value=Pressure)
                    for GasSequence, Gas in enumerate(GasType):
                        worksheet.cell(
                            row=2,
                            column=2 +
                            PressureSequence *
                            GasNum +
                            GasSequence,
                            value=Gas)

                for MaterialsSequence, Material in enumerate(
                        MaterialsNameList):
                    worksheet.cell(
                        row=3 + MaterialsSequence,
                        column=1,
                        value=Material)
                    for PressureSequence, Pressure in enumerate(PressureList):
                        for GasSequence, Gas in enumerate(GasType):
                            worksheet.cell(
                                row=3 + MaterialsSequence,
                                column=2 + PressureSequence * GasNum + GasSequence,
                                value=float(
                                    MaterialsListEnergyDictionary[Material][Partial][Temperature][Pressure][Gas]))

            Energy.save(
                os.path.join(
                    ExtractDataOutputPath,
                    '%s.xlsx' %
                    (Partial)))


def main():
    """main function"""

    print('Code is executing ......')
    ReadingPath, GasType, TemperatureList, PressureList, GasPartialPressure, ExtractEnergyData = ReadBasicInfo()
    MaterialsListLoadingDictionary, MaterialsListEnergyDictionary, MaterialsNameList = ExtractData(
        ReadingPath, GasType, TemperatureList, PressureList, ExtractEnergyData, GasPartialPressure)
    OutputTxt(
        ReadingPath,
        ExtractEnergyData,
        TemperatureList,
        PressureList,
        MaterialsNameList,
        GasType,
        MaterialsListLoadingDictionary,
        MaterialsListEnergyDictionary,
        GasPartialPressure)
    if OpenExcel:
        OutputExcel(
            ReadingPath,
            ExtractEnergyData,
            TemperatureList,
            PressureList,
            MaterialsNameList,
            GasType,
            MaterialsListLoadingDictionary,
            MaterialsListEnergyDictionary,
            GasPartialPressure)
    print("Completed!")


if __name__ == '__main__':
    main()
